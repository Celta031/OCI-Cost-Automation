"""Update an Excel workbook with monthly OCI costs by compartment."""

from __future__ import annotations

import argparse
import logging
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping

import openpyxl
from openpyxl.utils import column_index_from_string


LOGGER = logging.getLogger("oci_cost_updater")
ISO_UTC_FORMAT = "%Y-%m-%dT%H:%M:%SZ"
DEFAULT_COMPARTMENT_DEPTH = 6
DEFAULT_ENV_FILE = ".env"


class ConfigError(ValueError):
    """Raised when runtime configuration is invalid."""


@dataclass(frozen=True)
class AppConfig:
    excel_path: Path
    sheet_name: str | None
    start_row: int
    compartment_column: str
    target_column: str
    billing_start: str
    billing_end: str
    oci_profile: str | None = None
    compartment_depth: int = DEFAULT_COMPARTMENT_DEPTH
    dry_run: bool = False


@dataclass(frozen=True)
class ExcelUpdateResult:
    matched_compartments: list[str]
    updated_rows: int
    zeroed_rows: int
    skipped_rows: int
    saved: bool


def load_env_file(path: Path) -> dict[str, str]:
    """Load KEY=VALUE pairs from a dotenv-style file without extra dependencies."""
    if not path.exists():
        return {}

    values: dict[str, str] = {}
    for line_number, raw_line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            raise ConfigError(f"Invalid .env line {line_number}: expected KEY=VALUE")

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value

    return values


def env_value(env_file_values: Mapping[str, str], key: str) -> str | None:
    return os.environ.get(key) or env_file_values.get(key)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Update an Excel workbook with OCI costs grouped by compartment."
    )
    parser.add_argument("--excel-path", help="Path to the Excel workbook to update.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the active sheet when omitted.")
    parser.add_argument("--start-row", type=int, help="First data row in the worksheet.")
    parser.add_argument("--compartment-column", help="Column containing compartment names.")
    parser.add_argument("--target-column", help="Column where cost values will be written.")
    parser.add_argument("--start", help="Billing period start, ISO UTC format: YYYY-MM-DDTHH:MM:SSZ.")
    parser.add_argument("--end", help="Billing period end, ISO UTC format: YYYY-MM-DDTHH:MM:SSZ.")
    parser.add_argument("--oci-profile", help="OCI config profile name.")
    parser.add_argument("--compartment-depth", type=int, help="OCI compartment depth. Default: 6.")
    parser.add_argument("--env-file", default=DEFAULT_ENV_FILE, help="Path to .env file. Default: .env")
    parser.add_argument("--dry-run", action="store_true", help="Run all checks without saving the workbook.")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    return parser.parse_args(argv)


def build_config(args: argparse.Namespace) -> AppConfig:
    env_file_values = load_env_file(Path(args.env_file))

    def get(name: str, env_key: str) -> str | None:
        return getattr(args, name) or env_value(env_file_values, env_key)

    excel_path = get("excel_path", "OCI_EXCEL_PATH")
    sheet_name = get("sheet", "OCI_SHEET_NAME")
    start_row = args.start_row or env_value(env_file_values, "OCI_START_ROW")
    compartment_column = get("compartment_column", "OCI_COMPARTMENT_COLUMN")
    target_column = get("target_column", "OCI_TARGET_COLUMN")
    billing_start = get("start", "OCI_BILLING_START")
    billing_end = get("end", "OCI_BILLING_END")
    oci_profile = get("oci_profile", "OCI_PROFILE")
    compartment_depth = args.compartment_depth or env_value(env_file_values, "OCI_COMPARTMENT_DEPTH")

    missing = [
        label
        for label, value in {
            "--excel-path or OCI_EXCEL_PATH": excel_path,
            "--start-row or OCI_START_ROW": start_row,
            "--compartment-column or OCI_COMPARTMENT_COLUMN": compartment_column,
            "--target-column or OCI_TARGET_COLUMN": target_column,
            "--start or OCI_BILLING_START": billing_start,
            "--end or OCI_BILLING_END": billing_end,
        }.items()
        if value in (None, "")
    ]
    if missing:
        raise ConfigError("Missing required configuration: " + ", ".join(missing))

    config = AppConfig(
        excel_path=Path(str(excel_path)),
        sheet_name=sheet_name or None,
        start_row=int(str(start_row)),
        compartment_column=str(compartment_column),
        target_column=str(target_column),
        billing_start=str(billing_start),
        billing_end=str(billing_end),
        oci_profile=oci_profile or None,
        compartment_depth=int(str(compartment_depth or DEFAULT_COMPARTMENT_DEPTH)),
        dry_run=args.dry_run,
    )
    validate_config(config)
    return config


def validate_config(config: AppConfig) -> None:
    if config.start_row < 1:
        raise ConfigError("Start row must be greater than or equal to 1.")
    if config.compartment_depth < 1:
        raise ConfigError("Compartment depth must be greater than or equal to 1.")
    if not config.excel_path.exists():
        raise ConfigError(f"Excel file not found: {config.excel_path}")

    validate_column(config.compartment_column, "compartment column")
    validate_column(config.target_column, "target column")

    start = parse_utc_datetime(config.billing_start, "start")
    end = parse_utc_datetime(config.billing_end, "end")
    if start >= end:
        raise ConfigError("Billing start must be earlier than billing end.")


def validate_column(column: str, label: str) -> None:
    try:
        column_index_from_string(column.upper())
    except ValueError as exc:
        raise ConfigError(f"Invalid {label}: {column}") from exc


def parse_utc_datetime(value: str, label: str) -> datetime:
    try:
        return datetime.strptime(value, ISO_UTC_FORMAT)
    except ValueError as exc:
        raise ConfigError(
            f"Invalid {label} date: {value}. Expected format: YYYY-MM-DDTHH:MM:SSZ"
        ) from exc


def normalize_compartment_name(value: object) -> str:
    return str(value).strip()


def aggregate_cost_items(items: Iterable[object]) -> dict[str, float]:
    costs: dict[str, float] = {}
    for item in items:
        compartment_name = getattr(item, "compartment_name", None) or "Root"
        computed_amount = getattr(item, "computed_amount", None)
        amount = float(computed_amount if computed_amount is not None else 0.0)
        costs[compartment_name] = costs.get(compartment_name, 0.0) + amount
    return costs


def fetch_oci_costs(config: AppConfig) -> dict[str, float]:
    LOGGER.info("Connecting to OCI and requesting summarized costs.")
    try:
        import oci
    except ImportError as exc:
        raise RuntimeError("The 'oci' package is not installed. Run: pip install -r requirements.txt") from exc

    try:
        if config.oci_profile:
            oci_config = oci.config.from_file(profile_name=config.oci_profile)
        else:
            oci_config = oci.config.from_file()

        usage_client = oci.usage_api.UsageapiClient(oci_config)
        request_details = oci.usage_api.models.RequestSummarizedUsagesDetails(
            tenant_id=oci_config["tenancy"],
            time_usage_started=config.billing_start,
            time_usage_ended=config.billing_end,
            granularity="MONTHLY",
            query_type="COST",
            group_by=["compartmentName"],
            compartment_depth=config.compartment_depth,
        )
        response = usage_client.request_summarized_usages(request_details)
    except Exception as exc:
        raise RuntimeError(f"Failed to fetch OCI costs: {exc}") from exc

    costs = aggregate_cost_items(response.data.items)
    LOGGER.info("OCI returned costs for %s compartments.", len(costs))
    return costs


def update_excel(config: AppConfig, oci_costs: Mapping[str, float]) -> ExcelUpdateResult:
    LOGGER.info("Opening workbook: %s", config.excel_path)
    workbook = openpyxl.load_workbook(config.excel_path)

    if config.sheet_name:
        if config.sheet_name not in workbook.sheetnames:
            raise ConfigError(f"Worksheet not found: {config.sheet_name}")
        worksheet = workbook[config.sheet_name]
    else:
        worksheet = workbook.active

    matched: list[str] = []
    updated_rows = 0
    zeroed_rows = 0
    skipped_rows = 0

    for row in range(config.start_row, worksheet.max_row + 1):
        source_cell = worksheet[f"{config.compartment_column.upper()}{row}"]
        target_cell = worksheet[f"{config.target_column.upper()}{row}"]

        if source_cell.value in (None, ""):
            skipped_rows += 1
            continue

        compartment_name = normalize_compartment_name(source_cell.value)
        if compartment_name in oci_costs:
            target_cell.value = float(oci_costs[compartment_name])
            matched.append(compartment_name)
            updated_rows += 1
        else:
            target_cell.value = 0.0
            zeroed_rows += 1

    if config.dry_run:
        LOGGER.info("Dry-run enabled; workbook changes were not saved.")
        saved = False
    else:
        workbook.save(config.excel_path)
        LOGGER.info("Workbook saved successfully.")
        saved = True

    return ExcelUpdateResult(
        matched_compartments=matched,
        updated_rows=updated_rows,
        zeroed_rows=zeroed_rows,
        skipped_rows=skipped_rows,
        saved=saved,
    )


def find_unmapped_compartments(
    oci_costs: Mapping[str, float], matched_compartments: Iterable[str]
) -> dict[str, float]:
    matched = set(matched_compartments)
    return {name: oci_costs[name] for name in sorted(set(oci_costs) - matched)}


def log_report(oci_costs: Mapping[str, float], result: ExcelUpdateResult) -> None:
    unmapped = find_unmapped_compartments(oci_costs, result.matched_compartments)
    LOGGER.info("Rows updated with OCI costs: %s", result.updated_rows)
    LOGGER.info("Rows set to zero: %s", result.zeroed_rows)
    LOGGER.info("Rows skipped: %s", result.skipped_rows)
    LOGGER.info("Workbook saved: %s", "yes" if result.saved else "no")

    if unmapped:
        LOGGER.warning("OCI compartments with cost not found in the worksheet:")
        for name, value in unmapped.items():
            LOGGER.warning(" - %s: %.2f", name, value)
    else:
        LOGGER.info("All OCI compartments with cost were mapped in the worksheet.")


def configure_logging(level: str) -> None:
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    configure_logging(args.log_level)

    try:
        config = build_config(args)
        oci_costs = fetch_oci_costs(config)
        result = update_excel(config, oci_costs)
        log_report(oci_costs, result)
    except ConfigError as exc:
        LOGGER.error("%s", exc)
        return 2
    except Exception as exc:
        LOGGER.error("%s", exc)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
