from __future__ import annotations

import argparse
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from oci_cost_updater import (
    AppConfig,
    ConfigError,
    aggregate_cost_items,
    build_config,
    find_unmapped_compartments,
    normalize_compartment_name,
    parse_utc_datetime,
    update_excel,
)


def make_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "OCI CONSUMO"
    worksheet["A1"] = "Compartment"
    worksheet["A2"] = "ALG"
    worksheet["B3"] = "keeps max_row over a blank compartment row"
    worksheet["A4"] = "Finance"
    workbook.save(path)


def make_config(path: Path, dry_run: bool = False, sheet_name: str | None = "OCI CONSUMO") -> AppConfig:
    return AppConfig(
        excel_path=path,
        sheet_name=sheet_name,
        start_row=2,
        compartment_column="A",
        target_column="B",
        billing_start="2026-04-21T00:00:00Z",
        billing_end="2026-05-21T00:00:00Z",
        compartment_depth=6,
        dry_run=dry_run,
    )


def test_parse_utc_datetime_rejects_non_utc_format() -> None:
    with pytest.raises(ConfigError):
        parse_utc_datetime("2026-04-21", "start")


def test_normalize_compartment_name_strips_spaces() -> None:
    assert normalize_compartment_name("  Finance  ") == "Finance"


def test_aggregate_cost_items_sums_values_and_treats_none_as_zero() -> None:
    items = [
        SimpleNamespace(compartment_name="ALG", computed_amount=10.5),
        SimpleNamespace(compartment_name="ALG", computed_amount=None),
        SimpleNamespace(compartment_name=None, computed_amount=7),
    ]

    assert aggregate_cost_items(items) == {"ALG": 10.5, "Root": 7.0}


def test_find_unmapped_compartments_returns_oci_only_names() -> None:
    result = find_unmapped_compartments({"ALG": 10, "Finance": 5}, ["ALG"])

    assert result == {"Finance": 5}


def test_update_excel_writes_costs_and_zeroes_missing_compartments(tmp_path: Path) -> None:
    workbook_path = tmp_path / "costs.xlsx"
    make_workbook(workbook_path)

    result = update_excel(make_config(workbook_path), {"ALG": 42.25})

    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook["OCI CONSUMO"]
    assert worksheet["B2"].value == 42.25
    assert worksheet["B4"].value == 0
    assert result.updated_rows == 1
    assert result.zeroed_rows == 1
    assert result.skipped_rows == 1
    assert result.saved is True


def test_update_excel_dry_run_does_not_save(tmp_path: Path) -> None:
    workbook_path = tmp_path / "costs.xlsx"
    make_workbook(workbook_path)

    result = update_excel(make_config(workbook_path, dry_run=True), {"ALG": 42.25})

    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook["OCI CONSUMO"]
    assert worksheet["B2"].value is None
    assert result.saved is False


def test_update_excel_fails_when_sheet_does_not_exist(tmp_path: Path) -> None:
    workbook_path = tmp_path / "costs.xlsx"
    make_workbook(workbook_path)

    with pytest.raises(ConfigError, match="Worksheet not found"):
        update_excel(make_config(workbook_path, sheet_name="Missing"), {"ALG": 42.25})


def test_build_config_cli_overrides_env_file(tmp_path: Path) -> None:
    workbook_path = tmp_path / "costs.xlsx"
    make_workbook(workbook_path)
    env_file = tmp_path / ".env"
    env_file.write_text(
        "\n".join(
            [
                f"OCI_EXCEL_PATH={workbook_path}",
                "OCI_SHEET_NAME=OCI CONSUMO",
                "OCI_START_ROW=2",
                "OCI_COMPARTMENT_COLUMN=A",
                "OCI_TARGET_COLUMN=B",
                "OCI_BILLING_START=2026-01-01T00:00:00Z",
                "OCI_BILLING_END=2026-02-01T00:00:00Z",
            ]
        ),
        encoding="utf-8",
    )
    args = argparse.Namespace(
        excel_path=None,
        sheet=None,
        start_row=None,
        compartment_column=None,
        target_column="C",
        start="2026-04-21T00:00:00Z",
        end="2026-05-21T00:00:00Z",
        oci_profile=None,
        compartment_depth=None,
        env_file=str(env_file),
        dry_run=False,
    )

    config = build_config(args)

    assert config.target_column == "C"
    assert config.billing_start == "2026-04-21T00:00:00Z"
